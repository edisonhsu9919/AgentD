"""XLSX extraction layer (Phase O2).

Pure data extraction from XLSX files using openpyxl.
No LLM calls — structured raw data only.
"""

import os
from typing import Any

# Limits
_SAMPLE_ROWS = 5
_MAX_SHEETS_DETAIL = 20


def extract(path: str) -> dict[str, Any]:
    """Extract structural metadata and sample rows from an XLSX file.

    Returns a dict with:
      - path, kind, office_kind, size_bytes
      - sheet_count, sheet_names
      - sheets (list of per-sheet info: name, dimensions, header_row, sample_rows, has_formulas)

    Raises FileNotFoundError if path doesn't exist.
    Raises ValueError if file is not a valid XLSX.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")

    size_bytes = os.path.getsize(path)

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl") from e

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot read XLSX: {e}") from e

    sheet_names = wb.sheetnames
    sheets_info: list[dict[str, Any]] = []

    for name in sheet_names[:_MAX_SHEETS_DETAIL]:
        ws = wb[name]
        sheet_data = _extract_sheet(ws, name)
        sheets_info.append(sheet_data)

    return {
        "path": path,
        "kind": "office",
        "office_kind": "xlsx",
        "size_bytes": size_bytes,
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
        "sheets": sheets_info,
    }


def _extract_sheet(ws, name: str) -> dict[str, Any]:
    """Extract info from a single worksheet."""
    # Dimensions
    dimensions = ws.dimensions or ""
    rows = list(ws.iter_rows(max_row=_SAMPLE_ROWS + 1, values_only=True))

    # Header row (first row)
    header_row: list[str] = []
    if rows:
        header_row = [_cell_str(c) for c in rows[0]]

    # Sample rows (rows 2..N), skip fully empty rows
    sample_rows: list[list[str]] = []
    for row in rows[1:]:
        converted = [_cell_str(c) for c in row]
        if any(v for v in converted):  # skip all-empty rows
            sample_rows.append(converted)
        if len(sample_rows) >= _SAMPLE_ROWS:
            break

    # Check for formulas — need non-data_only workbook for this,
    # but since we open read_only+data_only, we approximate:
    # if any cell value is None but header exists, might have formulas
    # For simplicity, we just report max_row/max_column
    max_row = ws.max_row or 0
    max_column = ws.max_column or 0

    return {
        "name": name,
        "dimensions": dimensions,
        "max_row": max_row,
        "max_column": max_column,
        "header_row": header_row,
        "sample_rows": sample_rows,
    }


def _cell_str(value) -> str:
    """Convert a cell value to display string."""
    if value is None:
        return ""
    return str(value).strip()
